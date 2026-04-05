"""
excel.py — generates a formatted Excel report from a PortfolioSummary.
"""

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from config.settings import COLORS, CURRENCY_SYMBOL, EXCEL_SHEET_NAME, EXPORTS_DIR
from core.models import PortfolioSummary


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(hex_color: str = "000000", bold: bool = False, size: int = 11) -> Font:
    return Font(color=hex_color, bold=bold, size=size)


def generate(summary: PortfolioSummary, output_path: Path | None = None) -> Path:
    """
    Write a formatted Excel report and return the output path.
    If output_path is None, saves to data/exports/ with a timestamp filename.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = EXCEL_SHEET_NAME

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    ws["A1"] = f"Portfolio Report — {summary.generated_at.strftime('%d %b %Y, %I:%M %p')}"
    ws["A1"].font      = _font(COLORS["title_fg"], bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    # ── Headers ────────────────────────────────────────────────────────────────
    headers = [
        "Stock / ETF", "Ticker", "Category", "Qty",
        f"Avg Buy Price ({CURRENCY_SYMBOL})", f"Market Price ({CURRENCY_SYMBOL})",
        f"Invested ({CURRENCY_SYMBOL})", f"Market Value ({CURRENCY_SYMBOL})",
        f"P&L ({CURRENCY_SYMBOL})", "P&L (%)", "Status",
    ]
    ws.append([])
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.fill      = _fill(COLORS["header_bg"])
        cell.font      = _font(COLORS["header_fg"], bold=True)
        cell.alignment = Alignment(horizontal="center")

    # ── Data rows ──────────────────────────────────────────────────────────────
    def _write_section(holdings, section_label: str):
        if not holdings:
            return
        ws.append([section_label])
        label_row = ws.max_row
        ws.cell(label_row, 1).font = _font(bold=True, size=11)

        for h in holdings:
            status = "N/A"
            if h.pnl is not None:
                status = "▲ Profit" if h.is_profit else "▼ Loss"

            row = [
                h.name, h.ticker, h.category.upper(), h.qty,
                h.avg_price,
                h.market_price if h.market_price is not None else "ERROR",
                h.invested,
                h.market_value if h.market_value is not None else "N/A",
                h.pnl if h.pnl is not None else "N/A",
                f"{h.pnl_pct:+.2f}%" if h.pnl_pct is not None else "N/A",
                status,
            ]
            ws.append(row)
            r = ws.max_row
            if h.pnl is not None:
                bg = COLORS["profit_bg"] if h.is_profit else COLORS["loss_bg"]
                for col in range(1, 12):
                    ws.cell(r, col).fill = _fill(bg)

    _write_section(summary.stocks, "── Stocks")
    _write_section(summary.etfs,   "── ETFs")

    # ── Summary rows ───────────────────────────────────────────────────────────
    ws.append([])
    pnl_arrow = "▲ Profit" if summary.total_pnl >= 0 else "▼ Loss"
    summary_rows = [
        ["TOTAL INVESTED", "", "", "", "", "", f"{CURRENCY_SYMBOL} {summary.total_invested:,.2f}"],
        ["TOTAL MARKET VALUE", "", "", "", "", "", "", f"{CURRENCY_SYMBOL} {summary.total_market_value:,.2f}"],
        ["TOTAL P&L", "", "", "", "", "", "", "",
         f"{CURRENCY_SYMBOL} {summary.total_pnl:+,.2f}",
         f"{summary.total_pnl_pct:+.2f}%",
         pnl_arrow],
    ]
    for srow in summary_rows:
        ws.append(srow)
        r = ws.max_row
        for col in range(1, 12):
            ws.cell(r, col).fill = _fill(COLORS["summary_bg"])
            ws.cell(r, col).font = _font(bold=True)

    # ── Column widths ──────────────────────────────────────────────────────────
    widths = [28, 18, 10, 6, 18, 18, 15, 18, 14, 10, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Save ───────────────────────────────────────────────────────────────────
    if output_path is None:
        ts = summary.generated_at.strftime("%Y%m%d_%H%M")
        output_path = EXPORTS_DIR / f"Portfolio_Report_{ts}.xlsx"

    wb.save(output_path)
    return output_path
